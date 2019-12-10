import plotly.graph_objects as go
import squarify
import openpyxl

data_names = []
data_values = []
data_color = []

name_value_dict = dict()
name_color_dict = dict()
try:
    wb = openpyxl.load_workbook('heat_map_task.xlsx')
    sheet = wb.active

    # print (sheet.max_row)
    # print (sheet.max_column)
    for value in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=3, values_only= True):
        # print(value)
        try:
            if value[0] is not None and value[1] is not None and value[2] is not None:
                # data_names.append(value[0])
                # data_values.append(value[1])
                name_value_dict[value[0]] = value[1]
                name_color_dict[value[0]] = value[2]
        except Exception as e:
            continue
    sorted_name_value_dict = {k: v for k, v in sorted(name_value_dict.items(), key=lambda item: item[1], reverse=True)}
    for key,val in sorted_name_value_dict.items():
        data_names.append(key)
        data_values.append(val)
        # data_color.append(name_color_dict(key))
        color_name = name_color_dict.get(key)
        if  color_name == 'Dark Green':
            data_color.append('rgb(0,100,0)')
        elif color_name == 'One level below dark green':
            data_color.append('rgb(34,139,34)')
        elif color_name == 'Above Light Green':
            data_color.append('rgb(0,200,0)')
        elif color_name == 'Light green':
            data_color.append('rgb(144,238,144)')
        else:
            pass


    # print(data_names, data_values, data_color)
    fig = go.Figure()
    x = 0.
    y = 0.
    width = 50.
    height = 30.
    # width = 15.
    # height = 15.
    normed = squarify.normalize_sizes(data_values , width , height)
    rects = squarify.squarify(normed,x,y,width,height)
    color_brewer = data_color
    shapes=[]
    annotations = []
    counter=0

    pos=0
    for r, val, color in zip(rects, data_values, color_brewer):
        shapes.append(
            dict(
                type = 'rect',
                x0 = r['x'],
                y0 = r['y'],
                x1 = r['x']+r['dx'],
                y1 = r['y']+r['dy'],
                # x1 = r['x'],
                # y1 = r['y'],
                # x0 = r['x']+r['dx'],
                # y0 = r['y']+r['dy'],
                line = dict( width = 1 ),
                fillcolor = color
            )
        )
        data_name = (str(data_names[pos]).replace(' ','\n')+'('+str(val)+')')
        annotations.append(
            dict(
                x = r['x']+(r['dx']/2),
                y = r['y']+(r['dy']/2),
                # x = r['dx']+(r['x']/2),
                # y = r['dy']+(r['y']/2),
                text = data_name,
                showarrow = False
            )
        )
        pos += 1

    # For hover text
    fig.add_trace(go.Scatter(#x = [ r['x']+(r['dx']/2) for r in rects ],
                             #y = [ r['y']+(r['dy']/2) for r in rects ],
        x=[r['dx'] + (r['x'] / 2) for r in rects],
        y=[r['dy'] + (r['y'] / 2) for r in rects],
                             text = [ str(v) for v in data_values ],
                             mode = 'text',))

    fig.update_traces(textposition='top left')

    fig.update_layout(height=650,
                      width=1300,
                      xaxis=dict(showgrid=False,zeroline=False),
                      yaxis=dict(showgrid=False,zeroline=False),
                      shapes=shapes,
                      annotations=annotations,
                      font=dict(
                          # family="Times New Roman, BOLD",
                          size=11,
                          color="#FFFFFF"
                      ),
                      hovermode='closest')
    fig.show()
except FileNotFoundError as e:
    print(e)
    print('Seraching File Not Found')
except Exception as e:
    print(e)
    print('Exception Raised')