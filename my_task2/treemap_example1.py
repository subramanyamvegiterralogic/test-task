import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl

data_names = []
data_values = []
data_color = []

name_value_dict = dict()
name_color_dict = dict()
try:
    wb = openpyxl.load_workbook('heat_map_task.xlsx')
    sheet = wb.active

    # print (sheet.max_row)
    # print (sheet.max_column)
    for value in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=3, values_only= True):
        # print(value)
        try:
            if value[0] is not None and value[1] is not None and value[2] is not None:
                # data_names.append(value[0])
                # data_values.append(value[1])
                name_value_dict[value[0]] = value[1]
                name_color_dict[value[0]] = value[2]
        except Exception as e:
            continue
    sorted_name_value_dict = {k: v for k, v in sorted(name_value_dict.items(), key=lambda item: item[1], reverse=True)}
    for key,val in sorted_name_value_dict.items():
        data_names.append(key)
        data_values.append(val)
        # data_color.append(name_color_dict(key))
        color_name = name_color_dict.get(key)
        if  color_name == 'Dark Green':
            data_color.append('rgb(0,150,0)')
        elif color_name == 'One level below dark green':
            data_color.append('rgb(0,225,0)')
        elif color_name == 'Above Light Green':
            data_color.append('rgb(150,255,150)')
        elif color_name == 'Light green':
            data_color.append('rgb(229,255,229)')
        else:
            pass

    # labels = ["Eve", "Cain", "Seth", "Enos", "Noam", "Abel", "Awan", "Enoch", "Azura"]
    # parents = ["", "Eve", "Eve", "Seth", "Seth", "Eve", "Eve", "Awan", "Eve"]
    labels = data_names
    parents = data_names
    fig = make_subplots(
        cols=2, rows=1,
        column_widths=[0.4, 0.4],
        specs=[[{'type': 'treemap', 'rowspan': 1}, {'type': 'treemap'}]]
    )
    fig.add_trace(go.Treemap(
        ids=labels,
        labels=labels,
        parents=parents,
        # values=[10, 14, 12, 10, 2, 6, 6, 1, 4],
    ), row=1, col=1)

    fig.update_layout(
        margin={'t': 0, 'l': 0, 'r': 0, 'b': 0}
    )

    fig.show()
except Exception as e:
    print(e)
