import matplotlib.pyplot as plt
import squarify
import openpyxl

data_names = []
data_values = []
data_color = []
try:
    wb = openpyxl.load_workbook('heat_map_task.xlsx')
    sheet = wb.active

    # print (sheet.max_row)
    # print (sheet.max_column)
    for value in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=3, values_only= True):
        # print(value)
        try:
            if value[0] is not None and value[1] is not None and value[2] is not None:
                # data_names.append(value[0])
                data_names.append(str(value[0])+'({})'.format(str(value[1])))
                data_values.append(value[1])
                if value[2] == 'Dark Green':
                    data_color.append('darkgreen')
                elif value[2] == 'One level below dark green':
                    data_color.append('forestgreen')
                elif value[2] == 'Above Light Green':
                    data_color.append('limegreen')
                elif value[2] == 'Light green':
                    data_color.append('lightgreen')
                else:
                    pass
        except Exception as e:
            continue

    fig = plt.gcf()
    ax = fig.add_subplot()
    fig.set_size_inches(15, 15)
    shapes = []
    squarify.plot(sizes=data_values,
                  label=data_names,
                  color=data_color,
                  ax = ax,
                  shapes = shapes,
                  text_kwargs={'fontsize':10, 'fontname':"Times New Roman"},
                  bar_kwargs={'alpha':.9})

    # plt.title('Technical Strengths')
    # plt.axis('off')
    plt.show()

except FileNotFoundError:
    print(e)
    print('Seraching File Not Found')
except Exception as e:
    print(e)
    print('Exception Raised')